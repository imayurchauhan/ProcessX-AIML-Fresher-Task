"""
Compare our engine output against Sample_Input_Output.xlsx ground truth.
Run: .venv/Scripts/python.exe -m backend.app.test_sample --llm
     .venv/Scripts/python.exe -m backend.app.test_sample
"""
from __future__ import annotations
import asyncio, sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
SAMPLE = ROOT / 'Sample_Input_Output.xlsx'


def load_expected(wb, resident: str) -> list[dict]:
    """Load expected flags from the Output sheet (rows 4+)."""
    ws = wb[f'{resident} - Output']
    expected = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        day, flag, field, explanation = (row[i] if i < len(row) else None for i in range(4))
        if day is None and field is None:
            continue
        expected.append({'day': str(day or ''), 'flag_type': str(flag or ''), 'field': str(field or '')})
    return expected


def load_input_notes(wb, resident: str) -> list[dict]:
    ws = wb[f'{resident} - Input']
    days = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        day, date, staff, note = (row[i] if i < len(row) else None for i in range(4))
        if day is None and note is None:
            continue
        days.append({'day': str(day or '').strip(), 'note': str(note or '').strip()})
    return days


def normalise_field(f: str) -> str:
    return f.lower().strip().rstrip('.')


async def run_test(use_llm: bool) -> None:
    from backend.app.evaluator import evaluate_resident_notes_async, _evaluate_day_sync, _findings_to_issues
    from backend.app.llm import evaluate_note_with_groq, groq_available

    wb = load_workbook(SAMPLE)
    residents = ['John Doe', 'Peter Parker']

    engine_label = 'Groq LLM' if (use_llm and groq_available()) else 'Rule-based'
    print(f'\n{"="*65}')
    print(f'  Sample Accuracy Test  |  Engine: {engine_label}')
    print(f'{"="*65}')

    total_tp = total_fp = total_fn = 0

    for resident in residents:
        expected = load_expected(wb, resident)
        days = load_input_notes(wb, resident)

        # get our output
        issues, used_engine = await evaluate_resident_notes_async(days, resident)

        our_flags = {normalise_field(i.requirement) for i in issues if 'complete' not in i.flag_type.lower()}
        exp_flags = {normalise_field(e['field']) for e in expected if 'complete' not in e['flag_type'].lower()}

        tp = our_flags & exp_flags
        fp = our_flags - exp_flags
        fn = exp_flags - our_flags

        total_tp += len(tp)
        total_fp += len(fp)
        total_fn += len(fn)

        status = 'PASS' if not fp and not fn else 'FAIL'
        print(f'\n[{status}] {resident}  (engine used: {used_engine})')
        print(f'  Expected {len(exp_flags)} flags | Got {len(our_flags)} flags')

        if tp:
            print(f'  Correctly caught ({len(tp)}):')
            for f in sorted(tp):
                print(f'    + {f}')
        if fp:
            print(f'  False positives — should NOT have flagged ({len(fp)}):')
            for f in sorted(fp):
                print(f'    - {f}')
        if fn:
            print(f'  Missed — should have flagged ({len(fn)}):')
            for f in sorted(fn):
                print(f'    ! {f}')

    precision = total_tp / (total_tp + total_fp) if (total_tp + total_fp) > 0 else 1.0
    recall    = total_tp / (total_tp + total_fn) if (total_tp + total_fn) > 0 else 1.0
    f1        = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0.0

    print(f'\n{"-"*65}')
    print(f'  Precision : {precision:.1%}  (no false alarms)')
    print(f'  Recall    : {recall:.1%}  (caught all real issues)')
    print(f'  F1 Score  : {f1:.1%}')
    print(f'{"="*65}\n')


if __name__ == '__main__':
    use_llm = '--llm' in sys.argv
    asyncio.run(run_test(use_llm))
