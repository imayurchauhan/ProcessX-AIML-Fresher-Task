from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from .evaluator import Issue
from .excel_parser import parse_workbook as _parse

# ── colours matching the screenshots ─────────────────────────────────────────
_DARK_BLUE   = '1F4E79'   # title row bg
_MED_BLUE    = '2E75B6'   # column-header row bg
_LIGHT_GREEN = 'E2EFDA'   # complete-row bg
_WHITE       = 'FFFFFF'
_LIGHT_GREY  = 'F2F2F2'   # odd data rows

_THIN = Side(style='thin', color='BFBFBF')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_WRAP = Alignment(wrap_text=True, vertical='top')
_CENTER = Alignment(horizontal='center', vertical='top', wrap_text=True)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill('solid', fgColor=hex_color)


def _font(bold=False, color=_WHITE, size=11) -> Font:
    return Font(name='Arial', bold=bold, color=color, size=size)


def _write_input_sheet(ws, resident: object) -> None:
    """Re-style the Input sheet to match the screenshot theme."""
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 80

    # Row 1 – title
    ws.merge_cells('A1:D1')
    c = ws['A1']
    c.value = f"Progress Notes \u2014 {resident.resident_name}"
    c.font = _font(bold=True, size=13)
    c.fill = _fill(_DARK_BLUE)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    # Row 2 – subtitle
    ws.merge_cells('A2:D2')
    c = ws['A2']
    c.value = 'Fall incident: three consecutive daily progress notes.'
    c.font = Font(name='Arial', size=10, color='404040')
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 18

    # Row 3 – column headers
    headers = ['Day', 'Date', 'Documented By', 'Progress Note']
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = _font(bold=True, size=11)
        c.fill = _fill(_MED_BLUE)
        c.alignment = _CENTER
        c.border = _BORDER
    ws.row_dimensions[3].height = 20

    # Data rows (already in sheet – just re-style)
    for idx, day in enumerate(resident.days):
        row = 4 + idx
        fill = _fill(_LIGHT_GREY) if idx % 2 == 0 else _fill(_WHITE)
        for col, val in enumerate([day['day'], day.get('date', ''), day.get('staff_member', ''), day['note']], start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name='Arial', size=10, color='000000')
            c.fill = fill
            c.alignment = _WRAP if col == 4 else Alignment(horizontal='center', vertical='top', wrap_text=False)
            c.border = _BORDER
        ws.row_dimensions[row].height = 80


def _write_output_sheet(ws, resident_name: str, issues: list[Issue]) -> None:
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 70

    # Row 1 – title
    ws.merge_cells('A1:D1')
    c = ws['A1']
    c.value = f'Checker Output \u2014 {resident_name}'
    c.font = _font(bold=True, size=13)
    c.fill = _fill(_DARK_BLUE)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    # Row 2 – subtitle
    ws.merge_cells('A2:D2')
    c = ws['A2']
    c.value = 'All three days reviewed against the Falls Management Policy.'
    c.font = Font(name='Arial', size=10, color='404040')
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 18

    # Row 3 – column headers
    headers = ['Day', 'Flag Type', 'Field', 'Explanation']
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = _font(bold=True, size=11)
        c.fill = _fill(_MED_BLUE)
        c.alignment = _CENTER
        c.border = _BORDER
    ws.row_dimensions[3].height = 20

    for idx, issue in enumerate(issues):
        row = 4 + idx
        is_complete = 'Complete' in issue.flag_type or issue.flag_type == '✅ Complete'
        row_fill = _fill(_LIGHT_GREEN) if is_complete else (_fill(_LIGHT_GREY) if idx % 2 == 0 else _fill(_WHITE))
        for col, val in enumerate([issue.day, issue.flag_type, issue.requirement, issue.explanation], start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name='Arial', size=10, color='000000', bold=(col <= 2 and is_complete))
            c.fill = row_fill
            c.alignment = _WRAP if col == 4 else Alignment(horizontal='center' if col <= 2 else 'left', vertical='top', wrap_text=True)
            c.border = _BORDER
        ws.row_dimensions[row].height = 55


def write_issues_to_workbook(
    template_path: str | Path,
    output_path: str | Path,
    issues_by_resident: dict[str, list[Issue]],
) -> Path:
    source = load_workbook(template_path)
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    residents = {r.resident_name: r for r in _parse(template_path)}

    for sheet_name in source.sheetnames:
        is_input = sheet_name.endswith(' - Input')
        is_output = ('Output' in sheet_name) and ('Input' not in sheet_name)
        if not is_input and not is_output:
            continue

        resident_name = sheet_name.rsplit(' - ', 1)[0]
        ws = wb.create_sheet(title=sheet_name)

        if is_input and resident_name in residents:
            _write_input_sheet(ws, residents[resident_name])
        elif is_output:
            _write_output_sheet(ws, resident_name, issues_by_resident.get(resident_name, []))

    destination = Path(output_path)
    wb.save(destination)
    return destination
