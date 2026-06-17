from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import List

from openpyxl import load_workbook


@dataclass(frozen=True)
class ResidentSheet:
    resident_name: str
    input_sheet: str
    output_sheet: str
    days: list[dict]


def parse_workbook(path: str | Path) -> list[ResidentSheet]:
    workbook = load_workbook(path)
    residents: list[ResidentSheet] = []
    for sheet_name in workbook.sheetnames:
        if not sheet_name.endswith(' - Input'):
            continue
        resident_name = sheet_name.rsplit(' - ', 1)[0]
        output_sheet = next(
            (
                candidate
                for candidate in workbook.sheetnames
                if candidate.startswith(resident_name) and 'Output' in candidate and 'Input' not in candidate
            ),
            '',
        )
        worksheet = workbook[sheet_name]
        days: list[dict] = []
        for row_index in range(4, worksheet.max_row + 1):
            day = worksheet[f'A{row_index}'].value
            note = worksheet[f'D{row_index}'].value
            if day is None and note is None:
                continue
            days.append({
                'row': row_index,
                'day': str(day).strip() if day else f'Day {len(days) + 1}',
                'note': str(note or '').strip(),
                'date': worksheet[f'B{row_index}'].value,
                'staff_member': worksheet[f'C{row_index}'].value,
            })
        residents.append(ResidentSheet(resident_name=resident_name, input_sheet=sheet_name, output_sheet=output_sheet, days=days))
    return residents


def read_note_rows(path: str | Path) -> list[ResidentSheet]:
    return parse_workbook(path)


def read_resident_notes(workbook_path: str | Path) -> list[WorkbookResidentBundle]:
    workbook = load_workbook(workbook_path)
    bundles: list[WorkbookResidentBundle] = []
    input_sheets = [sheet for sheet in workbook.sheetnames if sheet.endswith(' - Input')]

    for input_sheet_name in input_sheets:
        resident = input_sheet_name.removesuffix(' - Input')
        output_sheet_name = f'{resident} - Output'
        worksheet = workbook[input_sheet_name]
        notes: list[ResidentDayNote] = []

        for row in range(4, worksheet.max_row + 1):
            day = _clean(worksheet[f'A{row}'].value)
            note = _clean(worksheet[f'D{row}'].value)
            if not day and not note:
                continue
            notes.append(
                ResidentDayNote(
                    resident=resident,
                    day=day or f'Day {len(notes) + 1}',
                    date=_clean(worksheet[f'B{row}'].value),
                    staff_member=_clean(worksheet[f'C{row}'].value),
                    note=note or '',
                )
            )

        bundles.append(
            WorkbookResidentBundle(
                resident=resident,
                input_sheet=input_sheet_name,
                output_sheet=output_sheet_name if output_sheet_name in workbook.sheetnames else None,
                notes=notes,
            )
        )

    return bundles
